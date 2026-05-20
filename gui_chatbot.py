import tkinter as tk
from tkinter import scrolledtext
import random
import json
import os
import re
import difflib  
import smtplib  
from email.mime.text import MIMEText  
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook, Workbook  


SENDER_EMAIL = "Enter_your_Store_Gmail_Here."      
SENDER_PASSWORD = "Enter_your_app_specific_Password_here."          

def send_real_gmail(receiver_email, subject, body_text):
    """Logs into real Google SMTP servers securely and dispatches an inbox notification."""
    if "your_store_email" in SENDER_EMAIL or len(SENDER_PASSWORD) < 10:
        print("⚠️ Email dispatch bypassed: Please configure real Gmail credentials at the top of the script.")
        return False
        
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()  
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        
        msg = MIMEMultipart()
        msg["From"] = f"Adarsh Garments <{SENDER_EMAIL}>"
        msg["To"] = receiver_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body_text, "plain"))
        
        server.sendmail(SENDER_EMAIL, receiver_email, msg.as_string())
        server.quit()
        print(f"📧 [REAL EMAIL SENT SUCCESSFULLY] Dispatched to: {receiver_email}")
        return True
    except Exception as e:
        print(f"⚠️ Failed to send real email via SMTP: {e}")
        return False


def initialize_excel_file():
    """Ensures orders.xlsx exists and sets up clean tabular headers if missing."""
    filename = "orders.xlsx"
    if not os.path.exists(filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Order Logs"
            headers = ["Order ID", "Status", "Customer Name", "Email", "Item / SKU", "Shipping Address", "System Notes"]
            ws.append(headers)
            wb.save(filename)
            print("📊 Initialized fresh Excel ledger spreadsheet: 'orders.xlsx'")
        except Exception as e:
            print(f"⚠️ Failed to create Excel file structure: {e}")

def log_order_to_excel(order_id, status, name, email, item, address, notes=""):
    """Appends a new structured transaction row into the Excel sheet."""
    filename = "orders.xlsx"
    initialize_excel_file()
    try:
        wb = load_workbook(filename)
        ws = wb.active
        row_data = [order_id, status, name, email, item, address, notes]
        ws.append(row_data)
        wb.save(filename)
    except Exception as e:
        print(f"⚠️ Error writing entry matrix rows to Excel: {e}")


def load_intents_json():
    """Loads intent patterns and responses directly from intents.json."""
    filename = "intents.json"
    if not os.path.exists(filename):
        print(f"⚠️ Warning: '{filename}' not found. Falling back to default responses.")
        return []
    try:
        with open(filename, "r", encoding="utf-8") as file:
            data = json.load(file)
            return data.get("intents", [])
    except Exception as e:
        print(f"⚠️ Error reading JSON file: {e}")
        return []

def get_current_offers():
    """Dynamically reads general promotions from current_offers.txt."""
    filename = "current_offers.txt"
    if not os.path.exists(filename):
        return "No special promotions running at the moment."
    try:
        with open(filename, "r", encoding="utf-8") as file:
            content = file.read().strip()
            return content if content else "No promotions listed today."
    except Exception as e:
        return "Unable to load current offers right now."

def get_live_discount():
    """Dynamically reads the active store discount from discount.txt on demand."""
    filename = "discount.txt"
    if not os.path.exists(filename):
        return "Get flat 10% off on your first order! Use code: ADARSH10"
    try:
        with open(filename, "r", encoding="utf-8") as file:
            content = file.read().strip()
            return content if content else "No active discounts available right now."
    except Exception as e:
        return "Error fetching latest discount details."

def generate_order_number():
    """Generates a unique 5-digit retail style Order ID prefixed with brand initials."""
    random_digits = random.randint(10000, 99999)
    return f"AG-{random_digits}"

def log_and_notify_order(order_id, order_details):
    """Logs order to Excel, alerts owner console, and sends real Confirmation Email."""
    log_order_to_excel(
        order_id=order_id, status="PLACED", name=order_details['name'],
        email=order_details['email'], item=order_details['item'],
        address=order_details['address'], notes="Order completed via GUI Checkout Workflow."
    )

    email_body = f"Hello {order_details['name']},\n\nThank you for shopping at Adarsh Garments!\n\nYour order has been confirmed successfully.\n📦 Item: {order_details['item']}\n🆔 Order Reference Number: {order_id}\n📍 Delivery Address: {order_details['address']}\n\nWe will email you the tracking link as soon as your package ships.\n\nBest Regards,\nAdarsh Garments Logistics Team"
    send_real_gmail(order_details['email'], f"Order Confirmed! Ticket: {order_id}", email_body)

    print("\n" + "="*45)
    print(f"🚨 [OWNER NOTIFICATION] NEW EXCEL ROW WRITTEN: {order_id} 🚨")
    print(f"Order ID:       {order_id}\nItem/SKU:       {order_details['item']}\nCustomer Name:  {order_details['name']}")
    print("="*45 + "\n")

def log_and_notify_cancellation(order_id, user_message):
    """Appends cancellation rows to Excel using fetched original user info and notifies both sides via email."""
    filename = "orders.xlsx"
    initialize_excel_file()
    
    
    name, email, item, address = "Unknown User", "N/A", "N/A", "N/A"
    
    try:
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == order_id:
                name = row[2]     
                email = row[3]    
                item = row[4]     
                address = row[5]  
                break             
    except Exception as e:
        print(f"⚠️ Could not read original order data for cancellation lookup: {e}")

   
    log_order_to_excel(
        order_id=order_id, status="CANCEL_REQ", name=name, email=email,
        item=item, address=address, notes=f"Reason: {user_message}"
    )

    
    if email != "N/A":
        customer_email_body = (
            f"Hello {name},\n\n"
            f"As requested, your order reference {order_id} has been successfully cancelled.\n\n"
            f"📦 Cancelled Item: {item}\n"
            f"💬 Reason Stated: {user_message}\n"
            f"💰 Refund Status: If your payment was already processed, a rollback will be initiated automatically within 2-3 business days.\n\n"
            f"Thank you for reaching out to us. If you didn't request this, please reply immediately.\n\n"
            f"Best Regards,\n"
            f"Adarsh Garments Support Team"
        )
        send_real_gmail(
            receiver_email=email, 
            subject=f"Order Cancelled Successfully - {order_id}", 
            body_text=customer_email_body
        )

    
    owner_email_body = (
        f"🚨 URGENT: ORDER CANCELLATION REQUESTED 🚨\n\n"
        f"Order Reference: {order_id}\n"
        f"Customer Name:   {name}\n"
        f"Customer Email:  {email}\n"
        f"Item/SKU ordered: {item}\n\n"
        f"User Message/Reason:\n'{user_message}'\n\n"
        f"Note: This transaction has been locked down as 'CANCEL_REQ' inside your orders.xlsx file."
    )
    send_real_gmail(
        receiver_email=SENDER_EMAIL, 
        subject=f"⚠️ ALERT: Cancellation Request for {order_id}", 
        body_text=owner_email_body
    )

    print("\n" + "🛑"*15)
    print(f"🚨 [OWNER NOTIFICATION] URGENT: ORDER CANCELLATION REQUESTED FOR {order_id}! 🚨")
    print(f"Customer Name:  {name} ({email})")
    print(f"User Reason: '{user_message}'")
    print("📧 Emails dispatched to both Customer and Owner consoles.")
    print("🛑"*15 + "\n")


class RetailChatbotGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Adarsh Garments AI Assistant")
        self.root.geometry("450x550")
        self.root.configure(bg="#f4f6f9")

        initialize_excel_file()
        self.intents_data = load_intents_json()

        self.order_state = None  
        self.current_order = {"item": "", "name": "", "email": "", "address": ""}
        
        
        self.ticket_action_mode = None 
        self.active_order_id = None      
        
        self.create_widgets()
        self.display_bot_message("Hello! Welcome to Adarsh Garments Menswear. How can I help you today?\n\n• Type 'offers' to see live deals.\n• Type 'discount' to get coupons.\n• Type 'order' to buy an item.")

    def create_widgets(self):
        self.chat_log = scrolledtext.ScrolledText(self.root, wrap=tk.WORD, state='disabled', bg="#ffffff", fg="#333333", font=("Arial", 10))
        self.chat_log.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        input_frame = tk.Frame(self.root, bg="#f4f6f9")
        input_frame.pack(padx=10, pady=(0, 10), fill=tk.X)

        self.user_input = tk.Entry(input_frame, font=("Arial", 11), bg="#ffffff")
        self.user_input.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)
        self.user_input.bind("<Return>", lambda event: self.handle_send())

        send_button = tk.Button(input_frame, text="Send", font=("Arial", 10, "bold"), bg="#007bff", fg="#ffffff", activebackground="#0056b3", activeforeground="#ffffff", relief=tk.FLAT, command=self.handle_send)
        send_button.pack(side=tk.RIGHT, padx=(5, 0), ipadx=10)

    def display_bot_message(self, text):
        self.chat_log.config(state='normal')
        self.chat_log.insert(tk.END, f"Bot: {text}\n\n")
        self.chat_log.config(state='disabled')
        self.chat_log.yview(tk.END)

    def handle_send(self):
        message = self.user_input.get().strip()
        if not message:
            return

        self.chat_log.config(state='normal')
        self.chat_log.insert(tk.END, f"You: {message}\n")
        self.chat_log.config(state='disabled')
        self.user_input.delete(0, tk.END)

        response = self.process_response(message)
        self.display_bot_message(response)

    def find_fuzzy_match(self, user_word, patterns_list):
        """NLP similarity engine to catch typos using Sequence Matching."""
        for pattern in patterns_list:
            pattern_clean = re.sub(r'[^\w\s]', '', pattern.lower())
            for p_word in pattern_clean.split():
                similarity = difflib.SequenceMatcher(None, user_word, p_word).ratio()
                if similarity > 0.82:  
                    return True
        return False

    def extract_order_id(self, text):
        """Extracts patterns like AG-12345 or simple 5 digit strings out of user messages."""
        match = re.search(r'(ag-\d{5})|(\b\d{5}\b)', text.lower())
        if match:
            return match.group(0).upper() if "ag-" in match.group(0) else f"AG-{match.group(0)}"
        return None

    def process_response(self, text):
        text_clean = re.sub(r'[^\w\s]', '', text.lower())
        words = text_clean.split()

        
        if self.ticket_action_mode == "WAITING_FOR_CANCEL_REASON":
            reason = text.strip()
            order_id = self.active_order_id
            
            
            self.ticket_action_mode = None
            self.active_order_id = None
            
            log_and_notify_cancellation(order_id, reason)
            return f"Thank you! 🛑\n\nYour cancellation has been registered for order {order_id} with the reason: \"{reason}\". Our backend system has been updated, and a confirmation mail has been sent to both you and our store manager."

        
        if self.ticket_action_mode in ["CANCEL", "EXCHANGE"]:
            extracted_id = self.extract_order_id(text)
            if not extracted_id:
                return "I couldn't find a valid Order Number. Order IDs are 5 digits long (e.g., AG-12345). Please type your correct Order ID to continue:"
            
            action = self.ticket_action_mode
            
            if action == "CANCEL":
                
                self.active_order_id = extracted_id
                self.ticket_action_mode = "WAITING_FOR_CANCEL_REASON"
                return f"Order {extracted_id} found! Please tell me your reason for cancelling this order so I can process your request:"
                
            elif action == "EXCHANGE":
                self.ticket_action_mode = None 
                name, email, item, address = "Unknown User", "N/A", "N/A", "N/A"
                try:
                    wb = load_workbook("orders.xlsx")
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row[0] == extracted_id:
                            name, email, item, address = row[2], row[3], row[4], row[5]
                            break
                except Exception as e:
                    print(f"⚠️ Excel read error during exchange lookup: {e}")

                log_order_to_excel(extracted_id, "EXCHANGE_REQ", name, email, item, address, "User requested exchange protocol loop.")
                return f"Thank you for providing your Order ID! 🔄\n\nYour Exchange Ticket has been processed for {name}. I have successfully initiated an exchange file under ID {extracted_id}. Our delivery team will contact you on your registered contact details within 24 hours to schedule the product swap."

        
        if self.order_state == "waiting_for_item":
            self.current_order["item"] = text
            self.order_state = "waiting_for_name"
            return "Got it! May I please have your full name for the delivery?"
        
        elif self.order_state == "waiting_for_name":
            self.current_order["name"] = text
            self.order_state = "waiting_for_email"
            return f"Thank you, {text}. What is your Gmail address? We will send your receipt and tracking details there."
        
        elif self.order_state == "waiting_for_email":
            self.current_order["email"] = text
            self.order_state = "waiting_for_address"
            return f"Thank you for providing your Gmail address! Finally, what is your complete shipping address?"
        
        elif self.order_state == "waiting_for_address":
            self.current_order["address"] = text
            self.order_state = None  
            
            new_id = generate_order_number()
            log_and_notify_order(new_id, self.current_order)
            return f"🎉 Success! Your order has been placed.\n\n🆔 ORDER NUMBER: {new_id}\n📦 Item: {self.current_order['item']}\n📧 A real confirmation mail is winging its way to {self.current_order['email']}.\n\nPlease note down your Order Number for future queries!"

        
        has_discount_keyword = any(self.find_fuzzy_match(w, ["discount", "discounts", "coupon", "coupons", "code"]) for w in words)
        if has_discount_keyword or "discount" in text_clean or "coupon" in text_clean:
            live_discount = get_live_discount()
            return f"🏷️ **Live Store Discount Details:**\n\n{live_discount}"

        has_offer_keyword = any(self.find_fuzzy_match(w, ["offers", "deals"]) for w in words)
        if has_offer_keyword or "offer" in text_clean or "deal" in text_clean:
            offers = get_current_offers()
            return f"🔥 Here are our current live offers:\n\n{offers}"
        
        
        has_cancel_keyword = any(self.find_fuzzy_match(w, ["cancel", "stop", "mistake"]) for w in words)
        if has_cancel_keyword and any(self.find_fuzzy_match(w, ["order", "shipping", "package"]) for w in words):
            self.ticket_action_mode = "CANCEL"
            return "I can certainly help you request a cancellation! To pull up your file details, could you please provide your 5-digit Order ID (e.g., AG-12345)?"

        
        has_exchange_keyword = any(self.find_fuzzy_match(w, ["exchange", "replace", "swap", "return"]) for w in words)
        if has_exchange_keyword and any(self.find_fuzzy_match(w, ["order", "size", "item", "shirt", "pant"]) for w in words):
            self.ticket_action_mode = "EXCHANGE"
            return "Changed your mind about the size or fit? No problem! Please enter your 5-digit Order ID so I can fetch your receipt profile:"

        
        has_order_keyword = any(self.find_fuzzy_match(w, ["order", "buy", "purchase"]) for w in words)
        if has_order_keyword:
            if not any(w in words for w in ["track", "status", "cancel", "where", "stop", "exchange", "return"]):
                self.order_state = "waiting_for_item"
                return "Let's get your order sorted out! What item or product SKU would you like to buy?"

        
        best_match_tag = None
        highest_score = 0
        matched_responses = []

        intent_weights = {
            "product_stock": 2.5, "track_order": 2.5, "payment_issues": 2.5,
            "cancel_order": 2.5, "returns_exchanges": 2.5, "greeting": 1.0, "goodbye": 1.0
        }

        user_words = set(words)

        for intent in self.intents_data:
            tag = intent.get("tag")
            weight = intent_weights.get(tag, 1.0)
            current_intent_responses = intent.get("responses", [])
            
            for pattern in intent.get("patterns", []):
                pattern_clean = re.sub(r'[^\w\s]', '', pattern.lower())
                pattern_words = pattern_clean.split()
                
                if not pattern_words:
                    continue
                
                if pattern_clean in text_clean:
                    exact_score = (len(pattern_words) * 3) * weight
                    if exact_score > highest_score:
                        highest_score = exact_score
                        best_match_tag = tag
                        matched_responses = current_intent_responses
                    continue
                
                matching_words_count = 0
                for p_word in pattern_words:
                    if p_word in user_words or self.find_fuzzy_match(p_word, words):
                        matching_words_count += 1
                
                if matching_words_count > 0:
                    match_ratio = matching_words_count / len(pattern_words)
                    if match_ratio >= 0.40:
                        calculated_score = (matching_words_count * 1.5) * weight * match_ratio
                        if calculated_score > highest_score:
                            highest_score = calculated_score
                            best_match_tag = tag
                            matched_responses = current_intent_responses

        print(f"\n--- DEBUG LOG ---")
        print(f"User Text: '{text}'")
        print(f"Highest Score calculated: {highest_score}")
        print(f"Best Match Tag found: '{best_match_tag}'")
        print(f"-----------------\n")

        if highest_score >= 1.5 and matched_responses:
            return random.choice(matched_responses)

        return "I'm sorry, I am a virtual assistant for Adarsh Garments and I can only help you with retail-related actions."


if __name__ == "__main__":
    root = tk.Tk()
    app = RetailChatbotGUI(root)
    root.mainloop()
